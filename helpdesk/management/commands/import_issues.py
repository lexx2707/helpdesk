from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from helpdesk.models import Category, IssueType

class Command(BaseCommand):
    help = "Import IssueType from Excel with columns: Category, Issue"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str)

    def handle(self, *args, **options):
        path = options["xlsx_path"]
        wb = load_workbook(path)
        ws = wb.active

        # หา header
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            idx_cat = headers.index("Category")
            idx_issue = headers.index("Issue")
        except ValueError:
            self.stderr.write("Excel ต้องมีคอลัมน์ชื่อ 'Category' และ 'Issue'")
            return

        created_cat = 0
        created_issue = 0
        updated_issue = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            cat_name = (row[idx_cat] or "")
            issue_name = (row[idx_issue] or "")
            cat_name = str(cat_name).strip()
            issue_name = str(issue_name).strip()

            if not cat_name or not issue_name:
                continue

            cat_obj, cat_created = Category.objects.get_or_create(name=cat_name)
            if cat_created:
                created_cat += 1

            obj, created = IssueType.objects.get_or_create(
                name=issue_name,
                category=cat_obj,
                defaults={"is_active": True},
            )
            if created:
                created_issue += 1
            else:
                # ถ้ามีอยู่แล้วให้ ensure active
                if not obj.is_active:
                    obj.is_active = True
                    obj.save(update_fields=["is_active"])
                    updated_issue += 1

        self.stdout.write(self.style.SUCCESS(
            f"Done. Category created: {created_cat}, Issue created: {created_issue}, Issue re-activated: {updated_issue}"
        ))
